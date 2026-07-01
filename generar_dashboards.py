#!/usr/bin/env python3
"""
generar_dashboards.py — ZAS Talents Dashboard Generator
Genera 6 HTMLs desde Odoo via JSON-RPC.
"""
import json, os, sys, urllib.request, http.cookiejar
from datetime import datetime, date, timezone
from collections import defaultdict, Counter
import openpyxl

# ─── CONFIG ────────────────────────────────────────────────────────────────────
ODOO_URL  = os.environ.get("ODOO_URL",  "https://zas-talent.odoo.com")
ODOO_DB   = os.environ.get("ODOO_DB",   "zas-talent")
ODOO_USER = os.environ.get("ODOO_USER", "martina.boazzo@zastalents.com")
ODOO_PASS = os.environ.get("ODOO_PASSWORD", "")
BUDGET_FILE = os.environ.get("BUDGET_FILE", "Base_Regional__2_.xlsx")
OBJETIVOS_FILE = os.environ.get("OBJETIVOS_FILE", "Objetivos_talentos_CLAUDE.xlsx")
OUTPUT_DIR  = os.environ.get("OUTPUT_DIR", "docs")

# Tipos de cambio a USD
def get_mep():
    """Obtiene el dólar MEP del día. Fallback: 1450."""
    try:
        req = urllib.request.Request(
            "https://dolarapi.com/v1/dolares/bolsa",
            headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            d = json.loads(r.read())
            if "venta" in d:
                return float(d["venta"])
    except Exception:
        pass
    try:
        req = urllib.request.Request(
            "https://api.bluelytics.com.ar/v2/latest",
            headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            d = json.loads(r.read())
            if "oficial" in d:
                return float(d["oficial"]["value_sell"])
    except Exception:
        pass
    return 1450.0

TC_ARS = get_mep()
TC = {'ARS': TC_ARS, 'CLP':940, 'COP':4250, 'PEN':3.72, 'USD':1, 'MXN':17}
TC_PRESUP_CLP = 890

# ─── CAMPOS A DESCARGAR ────────────────────────────────────────────────────────
TASK_FIELDS = [
    "id","name","company_id","sale_order_id","sale_line_id",
    "x_studio_fecha_de_publicacin",   # fecha publicación = implementado
    "x_studio_fecha_limite_ops",      # fecha estimada = pipeline
    "x_studio_related_field_7v0_1jidluau0",  # implementador
]

ORDER_FIELDS = [
    "id","name","state","company_id","partner_id","date_order",
    "x_studio_campaas_1",        # país de campaña (campo 1, USA lo usa)
    "x_studio_campaas",          # país de campaña (campo 2, fallback)
    "x_studio_bu_1",             # BU para internacionales
    "x_studio_tipo_de_contrato", # tipo: Comercial, Regional, Artistico, Canje, Agencia
]

LINE_FIELDS = [
    "id","order_id","task_id","price_unit","currency_id","state",
    "product_template_id",
]

# ─── CLASIFICACIÓN ─────────────────────────────────────────────────────────────
# x_studio_campaas_1 valor interno → (pais, tab)
CAMPAAS_1_MAP = {
    'Campañas':              ('argentina', 'local'),
    'Campañas Chile':        ('chile',     'local'),
    'Campañas Colombia':     ('colombia',  'local'),
    'Campañas Peru':         ('peru',      'local'),
    'Campañas USA':          ('usa',       'local'),
    'US Campaigns':          ('usa',       'local'),
    'Campañas Argentinas':   ('argentina', 'local'),
    'Campañas Colombia':     ('colombia',  'local'),
    'Campañas Mexico':       ('mexico',    'local'),   # nuevo dashboard México
    'Internacional':         (None,        'intl'),
    # Descartados: 'We Vibe','WOW',False
}

# x_studio_bu_1 → país para internacionales
BU_TO_PAIS = {
    'ZAS ARGENTINA': 'argentina',
    'ZAS CHILE':     'chile',
    'ZAS COLOMBIA':  'colombia',
    'ZAS USA':       'usa',
    'ZAS PERU':      'peru',
    'ZAS MEXICO':    'mexico',
    # Descartados: 'EXTERNOS','WE VIBE','WE ECHO'
}

PAIS_LABEL = {
    'argentina':    'Campañas Argentinas',
    'chile':        'Campañas Chilenas',
    'colombia':     'Campañas Colombianas',
    'usa':          'US Campaigns',
    'peru':         'Campañas Peruanas',
    'internacional':'Campañas Internacionales',
    'mexico':       'Campañas México',
}

# ─── ODOO JSON-RPC ─────────────────────────────────────────────────────────────
_cj = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cj))

def odoo_call(url, payload):
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data,
          headers={"Content-Type":"application/json"}, method="POST")
    with _opener.open(req, timeout=120) as r:
        return json.loads(r.read())

def odoo_auth():
    print("  Autenticando...")
    r = odoo_call(f"{ODOO_URL}/web/session/authenticate", {
        "jsonrpc":"2.0","method":"call","id":1,
        "params":{"db":ODOO_DB,"login":ODOO_USER,"password":ODOO_PASS}
    })
    uid = r["result"]["uid"]
    if not uid: raise Exception(f"Auth fallida: {r}")
    print(f"  ✓ UID={uid}")
    return uid

def search_read(model, domain, fields, batch=2000):
    ctx = {"allowed_company_ids":[1,2,3,4,5,6]}
    all_recs, offset = [], 0
    while True:
        r = odoo_call(f"{ODOO_URL}/web/dataset/call_kw", {
            "jsonrpc":"2.0","method":"call","id":2,
            "params":{"model":model,"method":"search_read","args":[domain],
                      "kwargs":{"fields":fields,"limit":batch,"offset":offset,"context":ctx}}
        })
        recs = r.get("result")
        if recs is None: raise Exception(f"Error {model}: {r.get('error')}")
        all_recs.extend(recs)
        print(f"    {model}: {len(all_recs)}...", end="\r")
        if len(recs) < batch: break
        offset += batch
    print(f"    {model}: {len(all_recs)} registros        ")
    return all_recs

# ─── DESCARGA ──────────────────────────────────────────────────────────────────
def download_data():
    print("  Descargando subtareas...")
    tasks = search_read("project.task",
        [["parent_id","!=",False]],
        TASK_FIELDS)

    print("  Descargando órdenes...")
    orders_raw = search_read("sale.order",
        [["state","in",["sale","done"]]],
        ORDER_FIELDS)

    print("  Descargando líneas...")
    lines_raw = search_read("sale.order.line",
        [["state","in",["sale","done"]]],
        LINE_FIELDS)

    # Índices
    orders = {o["id"]: o for o in orders_raw
              if (o.get("company_id")[0] if isinstance(o.get("company_id"),list) else o.get("company_id")) != 3}
    lines  = {l["id"]: l for l in lines_raw}

    print(f"  ✓ {len(tasks)} tareas, {len(orders)} órdenes, {len(lines)} líneas")

    return tasks, orders, lines

# ─── CLASIFICACIÓN ─────────────────────────────────────────────────────────────
def classify(tasks, orders):
    result = defaultdict(lambda: defaultdict(list))
    # Asegurar que México existe como key
    result['mexico']['local']  # inicializar
    stats = Counter()

    for t in tasks:
        cid = t.get("company_id")
        if isinstance(cid,list): cid=cid[0]
        # México (company_id=3) ya no se descarta — tiene su propio dashboard
        # Solo descartar si no tiene orden activa o no clasifica

        oid = t.get("sale_order_id")
        if isinstance(oid,list): oid=oid[0]
        order = orders.get(oid) if oid else None
        if not order:
            stats['sin_orden'] += 1
            continue

        t["_order"] = order

        v1 = order.get("x_studio_campaas_1") or ""
        v2 = order.get("x_studio_campaas")   or ""
        bu = (order.get("x_studio_bu_1")     or "").strip()

        # Usar v1 primero, v2 como fallback
        clasificacion = CAMPAAS_1_MAP.get(v1) or CAMPAAS_1_MAP.get(v2)

        if not clasificacion:
            stats[f'descartado:v1={v1}:v2={v2}'] += 1
            continue

        pais_raw, tab = clasificacion

        if tab == 'intl':
            # Solo cuando el campo dice 'Internacional' → derivar país por bu_1
            pais = BU_TO_PAIS.get(bu)
            if not pais:
                stats[f'intl_bu_desconocido:{bu}'] += 1
                continue
            result[pais]['intl'].append(t)
            result['internacional']['intl'].append(t)
            stats['intl'] += 1
        else:
            result[pais_raw]['local'].append(t)
            stats['local'] += 1

    print(f"  Clasificación:")
    for p in ['argentina','chile','colombia','usa','peru','internacional']:
        print(f"    {p}: {len(result[p]['local'])} local / {len(result[p]['intl'])} intl")
    print(f"  Stats: {dict(stats.most_common(5))}")
    return result

# ─── UTILIDADES ────────────────────────────────────────────────────────────────
def fmt_usd(v):
    if not v: return "—"
    v = float(v)
    if abs(v) >= 1_000_000: return f"USD {v/1_000_000:.1f}M"
    if abs(v) >= 1_000:     return f"USD {v/1_000:.1f}K"
    return f"USD {v:,.0f}"

def to_usd(amt, cur):
    if not amt: return 0.0
    return float(amt) / TC.get((cur or 'USD').upper(), 1)

def parse_date(s):
    if not s or s is False: return None
    try: return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except: return None

def get_linea(t, lines):
    slid = t.get("sale_line_id")
    if isinstance(slid,list): slid=slid[0]
    return lines.get(slid) if slid else None

def get_importe_usd(t, lines):
    linea = get_linea(t, lines)
    if linea:
        amt = linea.get("price_unit") or 0
        cur = linea.get("currency_id")
        cur_code = cur[1] if isinstance(cur,list) else "USD"
        return to_usd(float(amt), cur_code)
    return 0.0

def get_fecha_pub(t):
    return parse_date(t.get("x_studio_fecha_de_publicacin"))

def get_fecha_est(t):
    return parse_date(t.get("x_studio_fecha_limite_ops"))

def parse_talento(name, linea=None):
    # El nombre de la subtarea normalmente sigue el formato "Talento (Contenido)".
    # Cuando no lo sigue (nombres atípicos como "1 Reel", "Tik Tok"), la fuente
    # confiable es el producto vendido: en Odoo el producto ES el influencer y
    # los tipos de contenido están configurados como variantes, por lo que
    # product_template_id de la línea de venta trae el mismo formato
    # "Talento (Contenido)" pero siempre correcto.
    if "(" in name:
        tal = name[:name.index("(")].strip()
        if tal: return tal
    if linea:
        prod = linea.get("product_template_id")
        prod_name = prod[1] if isinstance(prod,list) and len(prod)>1 else ""
        if "(" in prod_name:
            tal = prod_name[:prod_name.index("(")].strip()
            if tal: return tal
        elif prod_name:
            return prod_name.strip()
    return None

def parse_contenido(name, linea=None):
    if "(" in name and ")" in name:
        return name[name.index("(")+1:name.rindex(")")].strip()
    if linea:
        prod = linea.get("product_template_id")
        prod_name = prod[1] if isinstance(prod,list) and len(prod)>1 else ""
        if "(" in prod_name and ")" in prod_name:
            return prod_name[prod_name.index("(")+1:prod_name.rindex(")")].strip()
    return ""

def pill_tipo(t):
    # Primero leer desde el campo de la orden (fuente confiable)
    order = t.get("_order")
    tipo_orden = (order.get("x_studio_tipo_de_contrato") or "") if order else ""
    if tipo_orden:
        mapa = {"Comercial":"pb","Regional":"pp","Artistico":"pg",
                "Artístico":"pg","Agencia":"pb","Canje":"pm2"}
        label_mapa = {"Comercial":"Comercial","Regional":"Regional",
                      "Artistico":"Artístico","Artístico":"Artístico",
                      "Agencia":"Comercial","Canje":"Canje"}
        if tipo_orden in mapa:
            return label_mapa[tipo_orden], mapa[tipo_orden]
    # Fallback por nombre de subtarea
    name = (t.get("name") or "").lower()
    cont = parse_contenido(t.get("name","")).lower()
    txt  = name + cont
    if any(k in txt for k in ["host","conducción","conduccion","exclusividad","artistico","artístico"]):
        return "Artístico","pg"
    if any(k in txt for k in ["regional","yerba"]):
        return "Regional","pp"
    if any(k in txt for k in ["canje","barter"]):
        return "Canje","pm2"
    return "Comercial","pb"

# ─── PRESUPUESTO ───────────────────────────────────────────────────────────────
BUDGET_CFG = {
    "argentina": {"sheet":"BG ARG","local":7,"comercial":18,"artistico":27,"regional":34,"intl":42,"tc":None},
    "chile":     {"sheet":"BG CHI","local":7,"comercial":17,"artistico":25,"regional":32,"intl":42,"tc":TC_PRESUP_CLP},
    "colombia":  {"sheet":"BG COL","local":7,"comercial":21,"artistico":30,"regional":37,"intl":46,"tc":None},
    "usa":       {"sheet":"BG USA","local":6,"comercial":17,"artistico":27,"regional":34,"intl":43,"tc":None},
    "peru":      {"sheet":None},
    "mexico":    {"sheet":None},
    "internacional":{"sheet":"BG INT","local":0},
}

def load_budget():
    if not os.path.exists(BUDGET_FILE):
        print(f"  ⚠ Budget file no encontrado: {BUDGET_FILE}")
        return {}
    wb = openpyxl.load_workbook(BUDGET_FILE, read_only=True, data_only=True)
    budgets = {}
    for pais, cfg in BUDGET_CFG.items():
        sname = cfg.get("sheet")
        if not sname or sname not in wb.sheetnames:
            budgets[pais] = {}
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        tc = cfg.get("tc")
        def get_row(idx):
            if idx is None or idx >= len(rows): return [0]*12
            row = rows[idx]
            vals = []
            for col in range(3,15):
                v = row[col] if col < len(row) else 0
                try: v = float(v or 0)
                except: v = 0.0
                if tc: v = v / tc
                vals.append(v)
            return vals
        budgets[pais] = {k: get_row(cfg.get(k)) for k in ["local","comercial","artistico","regional","intl"]}
    wb.close()
    return budgets

# ─── OBJETIVOS TALENTOS ─────────────────────────────────────────────────────────
# Código de país (columna B de la planilla) -> nombre interno usado en el dashboard
# (solo estos 4 tienen solapa propia en el HTML)
PAIS_CODE_MAP = {"AR":"argentina","CL":"chile","CO":"colombia","USA":"usa"}

# Columnas COMERCIAL/ARTISTICO por país en la planilla de objetivos (0-indexed),
# más el TC a aplicar (None = ya está en USD). Incluye TODOS los países del rango
# A-P (incluso MEX y PER, que no tienen solapa propia) porque participan en el
# cálculo de Intercompany de los demás.
OBJ_COLS_TODOS_PAISES = {
    # México: hoy no hay talentos con Comercial/Artístico cargado, así que no se
    # pudo confirmar si viene en millones de MXN (como Chile) o en unidades.
    # Se deja SIN escalar (tc=None) para evitar inflar por error si algún día
    # se carga un valor real — confirmar con Martina antes de cambiar esto.
    "mexico":    {"comercial":2,  "artistico":3,  "tc":None},
    "chile":     {"comercial":4,  "artistico":5,  "tc":TC_PRESUP_CLP},
    "peru":      {"comercial":6,  "artistico":7,  "tc":None},
    "argentina": {"comercial":8,  "artistico":9,  "tc":None},
    "usa":       {"comercial":10, "artistico":11, "tc":None},
    "colombia":  {"comercial":12, "artistico":13, "tc":None},
}
OBJ_COL_INTL = 14  # columna O — Internacional, USD, aplica a todos los talentos

def resolver_pais_talento(raw):
    """Resuelve el código de país crudo de la planilla de objetivos a AR/CL/CO/USA/etc.
    Reglas confirmadas con Martina:
    - 'NTW-XX' -> tomar el código después del guion
    - 'XX-YY' (dos países) -> tomar el primero
    - '-' o vacío -> sin país (se descarta)
    """
    if not raw or not str(raw).strip() or str(raw).strip() == "-":
        return None
    raw = str(raw).strip()
    if raw.upper().startswith("NTW"):
        partes = raw.split("-")
        return partes[1].strip() if len(partes) > 1 else None
    if "-" in raw:
        return raw.split("-")[0].strip()
    return raw

def _num(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return 0.0

def normalizar_nombre(n):
    """Normaliza nombres de talento para poder cruzar la planilla de objetivos
    (en MAYÚSCULAS) con los nombres reales de Odoo (capitalización normal)."""
    return " ".join((n or "").strip().upper().split())

def _valor_pais_usd(row, cfg):
    """Lee Comercial+Artístico de un país en la fila y los devuelve en USD."""
    comercial = _num(row[cfg["comercial"]]) if cfg["comercial"] < len(row) else 0.0
    artistico = _num(row[cfg["artistico"]]) if cfg["artistico"] < len(row) else 0.0
    if cfg["tc"]:
        # Chile/México vienen en millones de moneda local
        comercial = comercial * 1_000_000 / cfg["tc"]
        artistico = artistico * 1_000_000 / cfg["tc"]
    return comercial, artistico

def load_objetivos_talentos():
    """Lee la planilla de objetivos por talento (columnas A-P únicamente) y
    devuelve, por país interno (argentina/chile/colombia/usa), la lista de
    talentos con sus 4 objetivos en USD:
    - comercial: columna Comercial de su propio país
    - artistico: columna Artístico de su propio país
    - internacional: columna O, aplica a todos
    - intercompany: suma de Comercial+Artístico de TODOS LOS OTROS países
      (incluye México y Perú aunque no tengan solapa propia)
    """
    resultado = {p: [] for p in PAIS_CODE_MAP.values()}
    if not os.path.exists(OBJETIVOS_FILE):
        print(f"  ⚠ Archivo de objetivos no encontrado: {OBJETIVOS_FILE}")
        return resultado

    wb = openpyxl.load_workbook(OBJETIVOS_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    for row in rows[3:]:  # las primeras 3 filas son encabezados
        if not row or not row[0]:
            continue
        nombre = str(row[0]).strip()
        pais_raw = row[1] if len(row) > 1 else None
        pais_code = resolver_pais_talento(pais_raw)
        pais = PAIS_CODE_MAP.get(pais_code)
        if not pais:
            continue  # país no soportado en la nueva solapa (NUEVO sin país, etc.)

        comercial, artistico = _valor_pais_usd(row, OBJ_COLS_TODOS_PAISES[pais])

        internacional = _num(row[OBJ_COL_INTL]) if OBJ_COL_INTL < len(row) else 0.0

        intercompany = 0.0
        for otro_pais, cfg in OBJ_COLS_TODOS_PAISES.items():
            if otro_pais == pais:
                continue
            c, a = _valor_pais_usd(row, cfg)
            intercompany += c + a

        resultado[pais].append({
            "nombre": nombre,
            "nombre_norm": normalizar_nombre(nombre),
            "comercial": comercial,
            "artistico": artistico,
            "internacional": internacional,
            "intercompany": intercompany,
        })
    return resultado

TODOS_PAISES_LOCAL = ["argentina","chile","colombia","peru","usa","mexico"]

def compute_reales_talentos(pais, classified, lines, anio):
    """Para un país (argentina/chile/colombia/usa), calcula el acumulado real
    del año `anio` de cada talento que aparece en sus ventas locales o
    internacionales:
    - comercial: ventas locales del talento en su país, tipo Comercial
    - artistico: ventas locales del talento en su país, tipo Artístico
    - internacional: ventas del talento en 'Campañas Internacionales' (cualquier país destino)
    - intercompany: ventas del talento en el LOCAL de otros países, tipo Regional
    Solo cuenta tareas con fecha de publicación dentro de `anio`, ya que los
    objetivos de la planilla son anuales.
    Devuelve dict {nombre_talento_normalizado: {comercial, artistico, internacional, intercompany}}
    """
    reales = defaultdict(lambda: {"comercial":0.0,"artistico":0.0,"internacional":0.0,"intercompany":0.0})

    def es_del_anio(t):
        fpub = get_fecha_pub(t)
        return fpub is not None and fpub.year == anio

    # Comercial local + Artístico local (ventas locales de ESTE país)
    for t in classified[pais]['local']:
        if not es_del_anio(t): continue
        linea = get_linea(t, lines)
        tal = parse_talento(t.get("name",""), linea)
        if not tal: continue
        tal = normalizar_nombre(tal)
        tipo, _ = pill_tipo(t)
        amt = get_importe_usd(t, lines)
        if tipo == "Comercial":
            reales[tal]["comercial"] += amt
        elif tipo == "Artístico":
            reales[tal]["artistico"] += amt

    # Internacional (ventas del talento clasificadas como Campañas Internacionales,
    # sin importar a qué país se atribuyó por BU)
    for t in classified['internacional']['intl']:
        if not es_del_anio(t): continue
        linea = get_linea(t, lines)
        tal = parse_talento(t.get("name",""), linea)
        if not tal: continue
        tal = normalizar_nombre(tal)
        reales[tal]["internacional"] += get_importe_usd(t, lines)

    # Intercompany (ventas Regionales del talento en el LOCAL de OTROS países)
    for otro_pais in TODOS_PAISES_LOCAL:
        if otro_pais == pais: continue
        for t in classified[otro_pais]['local']:
            if not es_del_anio(t): continue
            linea = get_linea(t, lines)
            tal = parse_talento(t.get("name",""), linea)
            if not tal: continue
            tal = normalizar_nombre(tal)
            tipo, _ = pill_tipo(t)
            if tipo == "Regional":
                reales[tal]["intercompany"] += get_importe_usd(t, lines)

    return reales

# ─── KPIs ──────────────────────────────────────────────────────────────────────
def compute_kpis(tasks, lines, mes, anio):
    today = date.today()
    impl_mes = impl_ytd = 0.0
    cnt_mes = cnt_ytd = 0
    campanas_mes = set()
    pendientes = []
    vencidas = []
    talentos_mes = set()
    dias_cierre = []

    for t in tasks:
        fpub = get_fecha_pub(t)
        amt  = get_importe_usd(t, lines)
        order = t.get("_order")
        oid  = order["id"] if order else None

        if fpub:
            if fpub.year == anio and fpub.month == mes:
                impl_mes += amt; cnt_mes += 1
                if oid: campanas_mes.add(oid)
                tal = parse_talento(t.get("name",""), get_linea(t, lines))
                if tal: talentos_mes.add(tal)
            if fpub.year == anio and fpub.month <= mes:
                impl_ytd += amt; cnt_ytd += 1
            if order and order.get("date_order"):
                d0 = parse_date(order["date_order"])
                if d0:
                    dias = (fpub - d0).days
                    if 0 <= dias <= 365: dias_cierre.append(dias)
        else:
            pendientes.append(t)
            fest = get_fecha_est(t)
            if fest and fest < today: vencidas.append(t)

    pend_mes = [t for t in pendientes if get_fecha_est(t) and
                get_fecha_est(t).year==anio and get_fecha_est(t).month==mes]
    total_est = cnt_mes + len(pend_mes)
    tasa = cnt_mes / total_est if total_est else 0
    vel  = int(sum(dias_cierre)/len(dias_cierre)) if dias_cierre else 0
    sin_fecha = sum(1 for t in pendientes if not get_fecha_est(t))

    return {
        "impl_mes":impl_mes,"impl_ytd":impl_ytd,
        "cnt_mes":cnt_mes,"cnt_ytd":cnt_ytd,
        "campanas":len(campanas_mes),"talentos":len(talentos_mes),
        "tasa":tasa,"total_est":total_est,
        "vel":vel,"sin_fecha":sin_fecha,
        "pendientes":pendientes,"vencidas":vencidas,
    }

# ─── HTML ──────────────────────────────────────────────────────────────────────
CSS = """:root{--bg:#0d0f14;--s1:#161b24;--s2:#1e2535;--bdr:#262f42;
  --a1:#4f8ef7;--a2:#a78bfa;--a3:#34d399;--a4:#fb923c;--a5:#f472b6;
  --tx:#e8eaf0;--tx2:#8b93a8;--tx3:#5a6380;--red:#f87171;--yellow:#fbbf24}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--tx);font-family:'Inter',system-ui,sans-serif;font-size:13.5px;line-height:1.5}
.hdr{background:var(--s1);border-bottom:1px solid var(--bdr);padding:14px 28px;display:flex;align-items:center;gap:14px;position:sticky;top:0;z-index:200;flex-wrap:wrap}
.hdr-logo{font-size:18px;font-weight:700;color:var(--a1)}.hdr-sep{color:var(--tx3)}
.hdr-title{font-size:13px;color:var(--tx2);font-weight:500}
.hdr-badge{margin-left:auto;background:rgba(79,142,247,.12);color:var(--a1);border:1px solid rgba(79,142,247,.3);border-radius:20px;padding:3px 12px;font-size:11px;font-weight:600}
.hdr-date{font-size:11px;color:var(--tx3)}
.mtabs{display:flex;border-bottom:2px solid var(--bdr);padding:0 28px;background:var(--s1)}
.mtab{padding:14px 24px;cursor:pointer;font-size:14px;font-weight:600;color:var(--tx2);border-bottom:3px solid transparent;margin-bottom:-2px;transition:all .15s}
.mtab.local.active{color:var(--a1);border-bottom-color:var(--a1)}
.mtab.intl.active{color:var(--a2);border-bottom-color:var(--a2)}
.mtab.objetivos.active{color:var(--a4);border-bottom-color:var(--a4)}
.mpanel{display:none;padding:24px 28px}.mpanel.active{display:block}
.sec-label{display:inline-flex;align-items:center;gap:8px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;padding:5px 12px;border-radius:20px;margin-bottom:20px}
.sec-label.local{background:rgba(79,142,247,.1);color:var(--a1);border:1px solid rgba(79,142,247,.25)}
.sec-label.intl{background:rgba(167,139,250,.1);color:var(--a2);border:1px solid rgba(167,139,250,.25)}
.stabs{display:flex;gap:2px;border-bottom:1px solid var(--bdr);margin-bottom:24px;overflow-x:auto}
.stab{padding:10px 18px;cursor:pointer;font-size:12.5px;font-weight:500;color:var(--tx2);border-bottom:2px solid transparent;margin-bottom:-1px;white-space:nowrap;transition:all .15s}
.stab:hover{color:var(--tx)}.stab.active{color:var(--a1);border-bottom-color:var(--a1)}
.spanel{display:none}.spanel.active{display:block}
.badge-r{background:rgba(248,113,113,.2);color:var(--red);border-radius:10px;padding:1px 6px;font-size:10px}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:12px;margin-bottom:24px}
.kpi{background:var(--s1);border:1px solid var(--bdr);border-radius:10px;padding:16px 18px}
.kpi-l{font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:var(--tx3);margin-bottom:7px}
.kpi-v{font-size:22px;font-weight:700;letter-spacing:-.5px;line-height:1}
.kpi-s{font-size:11px;color:var(--tx3);margin-top:4px}
.kpi-n{font-size:10px;color:var(--tx3);margin-top:3px;font-style:italic;opacity:.8}
.c1{color:var(--a1)}.c2{color:var(--a2)}.c3{color:var(--a3)}.c4{color:var(--a4)}.cred{color:var(--red)}.cyellow{color:var(--yellow)}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:18px}
@media(max-width:900px){.g2{grid-template-columns:1fr}}
.card{background:var(--s1);border:1px solid var(--bdr);border-radius:10px;overflow:hidden;margin-bottom:18px}
.card-h{padding:12px 16px;border-bottom:1px solid var(--bdr);font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.7px;color:var(--tx2);display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.card-h-note{font-size:10px;color:var(--tx3);font-weight:400;text-transform:none;letter-spacing:0;margin-left:4px}
.dot{width:6px;height:6px;border-radius:50%;flex-shrink:0}
.card-b{padding:14px 16px}.note{font-size:11.5px;color:var(--tx3);padding:6px 0}
.obj-card summary::-webkit-details-marker{display:none}
.obj-card summary::marker{content:""}
.obj-card summary:hover{background:rgba(255,255,255,.02)}
.obj-card[open] summary{border-bottom:1px solid var(--bdr)}
.tw{overflow-x:auto}
table{width:100%;border-collapse:collapse}
th{text-align:left;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.6px;color:var(--tx3);padding:6px 9px;border-bottom:1px solid var(--bdr)}
th.r{text-align:right}
td{padding:7px 9px;border-bottom:1px solid #1a2030;font-size:12px;color:var(--tx)}
td.r{text-align:right;font-variant-numeric:tabular-nums}
td.m{color:var(--tx2)}
tr:last-child td{border-bottom:none}
tr:hover td{background:rgba(79,142,247,.04)}
.hbars{display:flex;flex-direction:column;gap:7px}
.hr{display:flex;align-items:center;gap:10px}
.hl{width:50px;font-size:11px;color:var(--tx2);text-align:right;flex-shrink:0}
.ht{flex:1;height:20px;background:var(--bdr);border-radius:3px;overflow:hidden}
.hf{height:100%;background:linear-gradient(90deg,var(--a1),var(--a2));border-radius:3px;display:flex;align-items:center;padding-left:8px}
.hv{font-size:10px;font-weight:600;color:rgba(255,255,255,.9)}
.pgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(105px,1fr));gap:10px;margin-bottom:20px}
.pm{background:var(--s2);border:1px solid var(--bdr);border-radius:10px;padding:14px 12px;text-align:center}
.pcur{border-color:var(--a1);background:rgba(79,142,247,.07)}
.pml{font-size:10px;text-transform:uppercase;letter-spacing:.7px;color:var(--tx3);margin-bottom:6px}
.pc{font-size:26px;font-weight:700;color:var(--a2);line-height:1;margin-bottom:3px}
.pi{font-size:10px;color:var(--tx2)}
.bgrow{display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--bdr)}
.bgrow:last-child{border-bottom:none}
.bgtipo{font-size:12px;font-weight:600;width:90px;flex-shrink:0}
.bgtrk{flex:1;height:11px;background:var(--bdr);border-radius:6px;overflow:hidden;position:relative}
.bgreal{height:100%;border-radius:6px;position:absolute;top:0;left:0}
.bgreal.ok{background:var(--a3)}.bgreal.warn{background:var(--yellow)}.bgreal.low{background:var(--red)}.bgreal.over{background:var(--a5)}
.bgnums{font-size:11px;color:var(--tx2);min-width:175px;text-align:right}
.bgpct{font-size:11px;font-weight:700;min-width:40px;text-align:right}
.acc{display:flex;flex-direction:column;gap:10px}
.ai{border-radius:9px;padding:13px 15px;display:flex;gap:11px;align-items:flex-start}
.aico{font-size:15px;flex-shrink:0;margin-top:2px}
.albl{font-size:10px;text-transform:uppercase;letter-spacing:.7px;font-weight:600;margin-bottom:4px}
.atxt{font-size:13px;line-height:1.6;color:var(--tx)}
.ai-urgente{background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.25)}.ai-urgente .albl{color:var(--red)}
.ai-alerta{background:rgba(251,191,36,.08);border:1px solid rgba(251,191,36,.25)}.ai-alerta .albl{color:var(--yellow)}
.ai-riesgo{background:rgba(251,146,60,.08);border:1px solid rgba(251,146,60,.25)}.ai-riesgo .albl{color:var(--a4)}
.ai-operativo{background:rgba(167,139,250,.08);border:1px solid rgba(167,139,250,.25)}.ai-operativo .albl{color:var(--a2)}
.ai-ok{background:rgba(52,211,153,.08);border:1px solid rgba(52,211,153,.25)}.ai-ok .albl{color:var(--a3)}
.pill{display:inline-block;padding:2px 8px;border-radius:12px;font-size:10px;font-weight:600}
.pb{background:rgba(79,142,247,.15);color:var(--a1)}
.pp{background:rgba(167,139,250,.15);color:var(--a2)}
.pg{background:rgba(52,211,153,.15);color:var(--a3)}
.pm2{background:rgba(90,99,128,.2);color:var(--tx2)}
.fxnote{background:rgba(167,139,250,.06);border:1px solid rgba(167,139,250,.2);border-radius:8px;padding:9px 13px;font-size:11px;color:var(--tx3);margin-bottom:18px}
.fxnote strong{color:var(--a2)}
.sec-title{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.8px;color:var(--tx3);margin-bottom:12px;padding-bottom:7px;border-bottom:1px solid var(--bdr)}"""

JS = """
function sw(pid,tab){
  var wrapper=document.getElementById('pg_'+pid);
  if(!wrapper)return;
  wrapper.querySelectorAll('.mtabs .mtab').forEach(t=>t.classList.remove('active'));
  wrapper.querySelectorAll('.mpanel').forEach(p=>p.classList.remove('active'));
  var el=wrapper.querySelector('.mtabs .mtab.'+tab);
  if(el)el.classList.add('active');
  var ep=document.getElementById(pid+'_'+tab);
  if(ep)ep.classList.add('active');
}
function sst(prefix,tab){
  var nav=document.getElementById(prefix+'_nav');
  if(!nav)return;
  nav.querySelectorAll('.stab').forEach(t=>t.classList.remove('active'));
  var pp=nav.parentElement;
  pp.querySelectorAll('.spanel').forEach(p=>p.classList.remove('active'));
  var trig=nav.querySelector('[onclick*=\\''+tab+'\\']');
  if(trig)trig.classList.add('active');
  var sp=document.getElementById(prefix+'_'+tab);
  if(sp)sp.classList.add('active');
}
function filtrarTalentos(inputId){
  var input=document.getElementById(inputId);
  var tabla=document.getElementById(inputId+'_list');
  if(!input||!tabla)return;
  var q=input.value.trim().toUpperCase();
  tabla.querySelectorAll('.obj-row').forEach(function(row){
    var nombre=row.getAttribute('data-nombre')||'';
    var visible = nombre.indexOf(q)!==-1;
    row.style.display = visible ? '' : 'none';
    var det=document.getElementById(row.getAttribute('data-rid'));
    if(det && !visible) det.style.display='none';
  });
}
function filtrarCategoria(inputId){
  var sel=document.getElementById(inputId+'_cat');
  var tabla=document.getElementById(inputId+'_list');
  if(!sel||!tabla)return;
  var vista=sel.value;
  tabla.querySelectorAll('.obj-vista').forEach(function(sp){
    sp.style.display = sp.getAttribute('data-vista')===vista ? '' : 'none';
  });
  ordenarPorVista(inputId, vista);
}
function ordenarPorVista(inputId, vista){
  var tabla=document.getElementById(inputId+'_list');
  if(!tabla)return;
  var tbody=tabla.tagName==='TBODY' ? tabla : tabla.querySelector('tbody');
  if(!tbody)return;
  var rows=Array.from(tbody.querySelectorAll('.obj-row'));
  var pares=rows.map(function(row){
    return {row:row, det:row.nextElementSibling, val:parseFloat(row.getAttribute('data-real-'+vista))||0};
  });
  pares.sort(function(a,b){ return b.val-a.val; });
  pares.forEach(function(p){
    tbody.appendChild(p.row);
    if(p.det) tbody.appendChild(p.det);
  });
}
function toggleObjDetalle(rid){
  var det=document.getElementById(rid);
  if(!det)return;
  det.style.display = det.style.display==='none' ? '' : 'none';
}"""

def mes_label(y,m):
    M=["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{M[m]} {str(y)[-2:]}"

def pct_cls(p):
    if p>1: return "over"
    if p>=.9: return "ok"
    if p>=.6: return "warn"
    return "low"

def pct_col(p):
    if p>1: return "var(--a5)"
    if p>=.9: return "var(--a3)"
    if p>=.6: return "var(--yellow)"
    return "var(--red)"

def build_implementado(tasks, lines, kpis, mes, anio):
    ml = mes_label(anio, mes)
    # Por tipo
    tipo_cnt = defaultdict(int); tipo_usd = defaultdict(float)
    # Por talento del mes
    tal_cnt = defaultdict(int); tal_usd = defaultdict(float)
    # Por cliente del mes
    cli_cnt = defaultdict(int); cli_usd = defaultdict(float)
    # Por implementador
    imp_cnt = defaultdict(int); imp_usd = defaultdict(float); imp_camp = defaultdict(set)

    for t in tasks:
        fpub = get_fecha_pub(t)
        if not (fpub and fpub.year==anio and fpub.month==mes): continue
        amt = get_importe_usd(t, lines)
        tipo, _ = pill_tipo(t)
        tipo_cnt[tipo]+=1; tipo_usd[tipo]+=amt
        tal = parse_talento(t.get("name",""), get_linea(t, lines))
        if tal:
            tal_cnt[tal]+=1; tal_usd[tal]+=amt
        order = t.get("_order")
        cli = ""
        if order and order.get("partner_id"):
            p = order["partner_id"]
            cli = p[1] if isinstance(p,list) else str(p)
        cli_cnt[cli]+=1; cli_usd[cli]+=amt
        imp_raw = t.get("x_studio_related_field_7v0_1jidluau0")
        imp_name = imp_raw[1] if isinstance(imp_raw,list) and len(imp_raw)>1 else "Sin asignar"
        imp_cnt[imp_name]+=1; imp_usd[imp_name]+=amt
        if order: imp_camp[imp_name].add(order["id"])

    def trow(tipo):
        if not tipo_cnt[tipo]: return ""
        _, pc = pill_tipo_by_name(tipo)
        return f'<tr><td><span class="pill {pc}">{tipo}</span></td><td class="r">{tipo_cnt[tipo]}</td><td class="r">{fmt_usd(tipo_usd[tipo])}</td></tr>'

    def pill_tipo_by_name(n):
        m={"Comercial":"pb","Regional":"pp","Artístico":"pg","Canje":"pm2"}
        return n, m.get(n,"pb")

    tipo_rows = "".join(trow(t) for t in ["Comercial","Regional","Artístico","Canje"])
    tal_rows  = "".join(f'<tr><td>{n}</td><td class="r">{tal_cnt[n]}</td><td class="r">{fmt_usd(tal_usd[n])}</td></tr>'
                        for n,_ in sorted(tal_usd.items(),key=lambda x:-x[1])[:20])
    cli_rows  = "".join(f'<tr><td>{n}</td><td class="r">{cli_cnt[n]}</td><td class="r">{fmt_usd(cli_usd[n])}</td></tr>'
                        for n,_ in sorted(cli_usd.items(),key=lambda x:-x[1])[:20])
    tot = kpis["impl_mes"] or 1
    conc_rows = "".join(
        f'<tr><td>{n}</td><td class="r">{cli_cnt[n]}</td><td class="r">{fmt_usd(cli_usd[n])}</td>'
        f'<td class="r" style="color:{pct_col(cli_usd[n]/tot)}">{cli_usd[n]/tot*100:.1f}%</td></tr>'
        for n,_ in sorted(cli_usd.items(),key=lambda x:-x[1])[:10])
    imp_rows = "".join(
        f'<tr><td>{n}</td><td class="r">{imp_cnt[n]}</td><td class="r">{len(imp_camp[n])}</td><td class="r">{fmt_usd(imp_usd[n])}</td></tr>'
        for n in sorted(imp_usd,key=lambda x:-imp_usd[x]))

    venc = kpis["vencidas"]
    usd_v = sum(get_importe_usd(t,lines) for t in venc)
    return f"""
<div class="kpis">
  <div class="kpi"><div class="kpi-l">Implementado {ml}</div><div class="kpi-v c1">{kpis['cnt_mes']}</div><div class="kpi-s">contenidos publicados</div></div>
  <div class="kpi"><div class="kpi-l">Importe</div><div class="kpi-v c3">{fmt_usd(kpis['impl_mes'])}</div><div class="kpi-s">en USD</div></div>
  <div class="kpi"><div class="kpi-l">Tasa Impl.</div><div class="kpi-v c3">{kpis['tasa']:.0%}</div><div class="kpi-s">{kpis['cnt_mes']} de {kpis['total_est']} est.</div><div class="kpi-n">% campañas estimadas efectivamente publicadas</div></div>
  <div class="kpi"><div class="kpi-l">Campañas Vencidas</div><div class="kpi-v cred">{len(venc)}</div><div class="kpi-s">{fmt_usd(usd_v)}</div><div class="kpi-n">Pendientes con fecha estimada ya pasada.</div></div>
  <div class="kpi"><div class="kpi-l">Pendientes</div><div class="kpi-v c4">{len(kpis['pendientes'])}</div><div class="kpi-s">{kpis['sin_fecha']} sin fecha</div></div>
  <div class="kpi"><div class="kpi-l">Vel. Cierre</div><div class="kpi-v c2">{kpis['vel']}</div><div class="kpi-s">días prom.</div><div class="kpi-n">Días desde pedido hasta publicación.</div></div>
</div>
<div class="g2">
  <div class="card"><div class="card-h"><div class="dot" style="background:var(--a2)"></div>Por Tipo</div>
    <div class="card-b"><div class="tw"><table><thead><tr><th>Tipo</th><th class="r">Cont.</th><th class="r">USD</th></tr></thead><tbody>{tipo_rows}</tbody></table></div></div></div>
  <div class="card"><div class="card-h"><div class="dot" style="background:var(--a3)"></div>Top Talentos del Mes</div>
    <div class="card-b"><div class="tw"><table><thead><tr><th>Talento</th><th class="r">Cont.</th><th class="r">USD</th></tr></thead><tbody>{tal_rows}</tbody></table></div></div></div>
</div>
<div class="card"><div class="card-h"><div class="dot" style="background:var(--a4)"></div>Top Clientes del Mes</div>
  <div class="card-b"><div class="tw"><table><thead><tr><th>Cliente</th><th class="r">Cont.</th><th class="r">USD</th></tr></thead><tbody>{cli_rows}</tbody></table></div></div></div>
<div class="card"><div class="card-h"><div class="dot" style="background:var(--a5)"></div>Concentración de Riesgo por Cliente</div>
  <div class="card-b"><div class="tw"><table><thead><tr><th>Cliente</th><th class="r">Cont.</th><th class="r">USD</th><th class="r">% total</th></tr></thead><tbody>{conc_rows}</tbody></table></div></div></div>
<div class="card"><div class="card-h"><div class="dot" style="background:var(--a1)"></div>Por Implementador</div>
  <div class="card-b"><div class="tw"><table><thead><tr><th>Implementador</th><th class="r">Contenidos</th><th class="r">Campañas</th><th class="r">USD</th></tr></thead><tbody>{imp_rows}</tbody></table></div></div></div>"""

def build_presupuesto(tasks, lines, budget_pais, mes, anio, tab):
    if not budget_pais: return '<p class="note">Presupuesto no disponible.</p>'
    ml = mes_label(anio, mes)
    mi = mes - 1

    def real_mes(tipo_check):
        total = 0.0
        for t in tasks:
            fpub = get_fecha_pub(t)
            if not (fpub and fpub.year==anio and fpub.month==mes): continue
            tipo, _ = pill_tipo(t)
            if tipo_check=="total" or tipo==tipo_check:
                total += get_importe_usd(t, lines)
        return total

    def real_ytd(tipo_check):
        total = 0.0
        for t in tasks:
            fpub = get_fecha_pub(t)
            if not (fpub and fpub.year==anio and fpub.month<=mes): continue
            tipo, _ = pill_tipo(t)
            if tipo_check=="total" or tipo==tipo_check:
                total += get_importe_usd(t, lines)
        return total

    filas = [("Total Local","total","local",True),("Comercial","Comercial","comercial",False),
             ("Artístico","Artístico","artistico",False),("Regional","Regional","regional",False)] if tab=="local" else \
            [("Internacional","total","intl",True)]

    html = f'<div class="card"><div class="card-h"><div class="dot" style="background:var(--a1)"></div>vs Presupuesto — Mes ({ml}) y YTD</div><div class="card-b"><p class="note" style="margin-bottom:14px">Verde ≥90% · Amarillo ≥60% · Rojo &lt;60% · Rosa = superado.</p>'

    for label, tipo_check, bkey, bold in filas:
        bdata = budget_pais.get(bkey, [0]*12)
        pm = bdata[mi] if mi < len(bdata) else 0
        py = sum(bdata[:mes])
        rm = real_mes(tipo_check)
        ry = real_ytd(tipo_check)
        if not pm: continue
        pct_m = rm/pm if pm else 0
        pct_y = ry/py if py else 0
        bw = min(pct_m*100, 100)
        style = "font-weight:700" if bold else ""
        html += f"""<div class="bgrow">
  <span class="bgtipo" style="width:100px;{style}">{label}</span>
  <div style="flex:1"><div style="margin-bottom:4px"><span style="font-size:10px;color:var(--tx3)">Mes ({ml})</span></div>
    <div class="bgtrk"><div class="bgreal {pct_cls(pct_m)}" style="width:{bw:.0f}%"></div></div></div>
  <div class="bgnums">{fmt_usd(rm)} / {fmt_usd(pm)}</div>
  <div class="bgpct" style="color:{pct_col(pct_m)}">{pct_m:.0%}</div>
</div>
<div class="bgrow">
  <span class="bgtipo" style="width:100px;color:var(--tx3);font-size:11px">{label} YTD</span>
  <div style="flex:1"><div style="margin-bottom:4px"><span style="font-size:10px;color:var(--tx3)">YTD</span></div>
    <div class="bgtrk"><div class="bgreal {pct_cls(pct_y)}" style="width:{min(pct_y*100,100):.0f}%"></div></div></div>
  <div class="bgnums">{fmt_usd(ry)} / {fmt_usd(py)}</div>
  <div class="bgpct" style="color:{pct_col(pct_y)}">{pct_y:.0%}</div>
</div>"""

    return html + "</div></div>"

def build_objetivos_talentos(pais, anio, classified, lines, objetivos_pais):
    """Arma la solapa 'Objetivos Talentos' como una única tabla: una fila por
    talento con Real/Objetivo/%/Falta según la categoría elegida en el filtro
    (Total por defecto). Al tocar el nombre se despliega una fila de detalle
    con el desglose completo de Comercial/Artístico/Internacional/Intercompany.
    Incluye buscador por nombre."""
    if not objetivos_pais:
        return '<p class="note">No hay objetivos de talentos cargados para este país.</p>'

    reales = compute_reales_talentos(pais, classified, lines, anio)

    conceptos = [
        ("total",         "Total"),
        ("comercial",     "Comercial"),
        ("artistico",     "Artístico"),
        ("internacional", "Internacional"),
        ("intercompany",  "Intercompany"),
    ]

    def celda_valores(real, objetivo):
        """Devuelve (real_str, objetivo_str, pct_str, pct_color, falta_str) para una categoría."""
        if not objetivo and not real:
            return "—", "Sin objetivo", "—", "var(--tx3)", "—"
        if not objetivo:
            return fmt_usd(real), "Sin objetivo", "—", "var(--tx3)", "Sin objetivo cargado"
        pct = real/objetivo
        falta = max(objetivo - real, 0)
        return fmt_usd(real), fmt_usd(objetivo), f"{pct:.0%}", pct_col(pct), (fmt_usd(falta) if falta else "✓ Cumplido")

    def fila_bg_desglose(label, real, objetivo):
        """Fila con barra de progreso para el desglose expandido."""
        if not objetivo and not real:
            return ""
        if not objetivo:
            return f"""<div class="bgrow">
  <span class="bgtipo" style="width:100px">{label}</span>
  <div style="flex:1"><div class="bgtrk"><div class="bgreal warn" style="width:100%"></div></div></div>
  <div class="bgnums">{fmt_usd(real)} / Sin objetivo</div>
  <div class="bgpct" style="color:var(--tx3)">—</div>
  <div class="bgnums" style="min-width:110px;color:var(--tx3)">Sin objetivo cargado</div>
</div>"""
        pct = real/objetivo
        falta = max(objetivo - real, 0)
        bw = min(pct*100, 100)
        return f"""<div class="bgrow">
  <span class="bgtipo" style="width:100px">{label}</span>
  <div style="flex:1"><div class="bgtrk"><div class="bgreal {pct_cls(pct)}" style="width:{bw:.0f}%"></div></div></div>
  <div class="bgnums">{fmt_usd(real)} / {fmt_usd(objetivo)}</div>
  <div class="bgpct" style="color:{pct_col(pct)}">{pct:.0%}</div>
  <div class="bgnums" style="min-width:110px;color:var(--tx3)">Falta: {fmt_usd(falta) if falta else "✓ Cumplido"}</div>
</div>"""

    def total_real_de(obj):
        r = reales.get(obj["nombre_norm"], {"comercial":0.0,"artistico":0.0,"internacional":0.0,"intercompany":0.0})
        return sum(r.values())

    filas_tabla = ""
    row_id = 0
    for obj in sorted(objetivos_pais, key=lambda o: -total_real_de(o)):
        nombre = obj["nombre"]
        r = reales.get(obj["nombre_norm"], {"comercial":0.0,"artistico":0.0,"internacional":0.0,"intercompany":0.0})

        desglose = "".join(fila_bg_desglose(label, r[key], obj[key]) for key, label in conceptos[1:])
        if not desglose:
            continue  # talento sin ningún dato (ni objetivo ni real) en las 4 categorías

        total_real = sum(r.values())
        total_obj  = sum(obj[k] for k,_ in conceptos[1:])

        # Una <td> por celda de valor, por cada vista posible; el JS oculta todas
        # menos la que corresponde a la categoría elegida en el <select>
        celdas_por_vista = {}
        for key, _ in conceptos:
            real, objetivo = (total_real, total_obj) if key == "total" else (r[key], obj[key])
            celdas_por_vista[key] = celda_valores(real, objetivo)

        def celda_vista(idx):
            spans = "".join(
                f'<span class="obj-vista" data-vista="{key}" style="{"" if key=="total" else "display:none"}">{celdas_por_vista[key][idx]}</span>'
                for key,_ in conceptos
            )
            return spans

        row_id += 1
        rid = f"objrow_{pais}_{row_id}"
        nombre_attr = normalizar_nombre(nombre)

        # Valor real crudo por vista, usado por el JS para reordenar la tabla
        # de mayor a menor según la categoría elegida en el <select>
        reales_attrs = " ".join(
            f'data-real-{key}="{(total_real if key=="total" else r[key]):.2f}"'
            for key,_ in conceptos
        )

        # Color dinámico del % (solo se ve el de la vista activa, pero por simplicidad
        # de markup mostramos el texto con el color correspondiente vía inline en cada span)
        pct_spans = "".join(
            f'<span class="obj-vista" data-vista="{key}" style="color:{celdas_por_vista[key][3]};{"" if key=="total" else "display:none"}">{celdas_por_vista[key][2]}</span>'
            for key,_ in conceptos
        )

        filas_tabla += f"""<tr class="obj-row" data-nombre="{nombre_attr}" data-rid="{rid}" {reales_attrs} onclick="toggleObjDetalle('{rid}')" style="cursor:pointer">
  <td style="font-weight:600">{nombre}</td>
  <td class="r">{celda_vista(0)}</td>
  <td class="r">{celda_vista(1)}</td>
  <td class="r" style="font-weight:700">{pct_spans}</td>
  <td class="r">{celda_vista(4)}</td>
</tr>
<tr class="obj-detalle" id="{rid}" style="display:none">
  <td colspan="5" style="padding:14px 16px;background:var(--bg)">{desglose}</td>
</tr>"""

    if not filas_tabla:
        return '<p class="note">No hay objetivos de talentos con datos cargados.</p>'

    buscador_id = f"objbuscar_{pais}"
    controles = f"""<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px;align-items:center">
  <input type="text" id="{buscador_id}" placeholder="🔎 Buscar talento..." autocomplete="off"
    oninput="filtrarTalentos('{buscador_id}')"
    style="flex:1;min-width:200px;max-width:320px;padding:8px 12px;border-radius:8px;border:1px solid var(--bdr);background:var(--s1);color:var(--tx1);font-size:13px;font-family:inherit">
  <select id="{buscador_id}_cat" onchange="filtrarCategoria('{buscador_id}')"
    style="padding:8px 12px;border-radius:8px;border:1px solid var(--bdr);background:var(--s1);color:var(--tx1);font-size:13px;font-family:inherit">
    <option value="total">Total</option>
    <option value="comercial">Comercial</option>
    <option value="artistico">Artístico</option>
    <option value="internacional">Internacional</option>
    <option value="intercompany">Intercompany</option>
  </select>
</div>"""

    tabla = f"""<div class="tw"><table id="{buscador_id}_list">
<thead><tr><th>Talento</th><th class="r">Real</th><th class="r">Objetivo</th><th class="r">%</th><th class="r">Falta</th></tr></thead>
<tbody>{filas_tabla}</tbody>
</table></div>"""

    return f'<p class="sec-title">Objetivos {anio} por Talento — Acumulado Real vs Objetivo</p>{controles}{tabla}'

def build_pipeline(tasks, lines, mes, anio):
    today = date.today()
    pendientes = [t for t in tasks if not get_fecha_pub(t)]

    by_mes = defaultdict(list)
    sin_fecha = []
    for t in pendientes:
        fe = get_fecha_est(t)
        if fe: by_mes[(fe.year,fe.month)].append(t)
        else:  sin_fecha.append(t)

    pgrid = ""
    for (y,m) in sorted(by_mes.keys())[-12:]:
        items = by_mes[(y,m)]
        usd = sum(get_importe_usd(t,lines) for t in items)
        cur = "pcur" if y==anio and m==mes else ""
        pgrid += f'<div class="pm {cur}"><div class="pml">{mes_label(y,m)}</div><div class="pc">{len(items)}</div><div class="pi">{fmt_usd(usd)}</div></div>'

    venc = sorted([t for t in pendientes if get_fecha_est(t) and get_fecha_est(t)<today],
                  key=lambda t: get_fecha_est(t))

    def trow(t, show_dias=True):
        nombre = t.get("name","")
        linea = get_linea(t, lines)
        tal = parse_talento(nombre, linea) or "—"
        cont = parse_contenido(nombre, linea)
        order = t.get("_order")
        cli = ""
        if order and order.get("partner_id"):
            p = order["partner_id"]
            cli = p[1] if isinstance(p,list) else ""
        tipo, pc = pill_tipo(t)
        usd = get_importe_usd(t,lines)
        fe = get_fecha_est(t)
        fe_str = fe.strftime("%d/%m/%Y") if fe else "—"
        dias_td = ""
        if show_dias and fe and fe < today:
            dias_td = f'<td class="r" style="color:var(--red)">{(today-fe).days}d</td>'
        elif show_dias:
            dias_td = '<td class="r">—</td>'
        return (f'<tr><td>{tal}</td><td class="m">{cont}</td><td class="m">{cli}</td>'
                f'<td><span class="pill {pc}">{tipo}</span></td>'
                f'<td class="r">{fmt_usd(usd)}</td><td class="m">{fe_str}</td>{dias_td}</tr>')

    venc_rows = "".join(trow(t) for t in venc[:50])
    venc_card = ""
    if venc:
        venc_card = f'<div class="card" style="margin-bottom:20px"><div class="card-h"><div class="dot" style="background:var(--red)"></div>Campañas Vencidas ({len(venc)})</div><div class="card-b"><div class="tw"><table><thead><tr><th>Talento</th><th>Contenido</th><th>Cliente</th><th>Tipo</th><th class=r>USD</th><th>F.Estimada</th><th class=r>Días</th></tr></thead><tbody>{venc_rows}</tbody></table></div></div></div>'

    top_rows = "".join(trow(t,False) for t in sorted(pendientes,key=lambda t:-get_importe_usd(t,lines))[:30])
    sf_note = f'<p class="note">+ {len(sin_fecha)} sin fecha estimada.</p>' if sin_fecha else ""

    return f'<p class="sec-title">Pipeline por Mes Estimado</p><div class="pgrid">{pgrid}</div>{sf_note}{venc_card}<p class="sec-title">Top Pendientes por Importe</p><div class="tw"><table><thead><tr><th>Talento</th><th>Contenido</th><th>Cliente</th><th>Tipo</th><th class=r>USD</th><th>F.Estimada</th></tr></thead><tbody>{top_rows}</tbody></table></div>'

def build_historial(tasks, lines, mes, anio):
    by_mes = defaultdict(lambda: {"cnt":0,"usd":0.0})
    for t in tasks:
        fp = get_fecha_pub(t)
        if not fp: continue
        key = (fp.year,fp.month)
        by_mes[key]["cnt"]+=1
        by_mes[key]["usd"]+=get_importe_usd(t,lines)

    meses = sorted(by_mes.keys())
    max_cnt = max((v["cnt"] for v in by_mes.values()),default=1) or 1
    hbars = ""
    for (y,m) in meses[-18:]:
        v = by_mes[(y,m)]
        pct = int(v["cnt"]/max_cnt*100)
        hbars += f'<div class="hr"><div class="hl">{mes_label(y,m)}</div><div class="ht"><div class="hf" style="width:{pct}%"><span class="hv">{v["cnt"]}</span></div></div><div style="font-size:10px;color:var(--tx3);white-space:nowrap;min-width:80px;text-align:right">{fmt_usd(v["usd"])}</div></div>'

    tal_cnt = defaultdict(int); tal_usd = defaultdict(float)
    for t in tasks:
        fp = get_fecha_pub(t)
        if not fp: continue
        tal = parse_talento(t.get("name",""), get_linea(t, lines))
        if not tal: continue
        tal_cnt[tal]+=1; tal_usd[tal]+=get_importe_usd(t,lines)
    top_tal = sorted(tal_usd.items(),key=lambda x:-x[1])[:15]
    tal_rows = "".join(f'<tr><td>{n}</td><td class="r">{tal_cnt[n]}</td><td class="r">{fmt_usd(u)}</td></tr>' for n,u in top_tal)

    cli_cnt = defaultdict(int); cli_usd = defaultdict(float)
    for t in tasks:
        fp = get_fecha_pub(t)
        if not fp: continue
        order = t.get("_order")
        cli = ""
        if order and order.get("partner_id"):
            p = order["partner_id"]
            cli = p[1] if isinstance(p,list) else ""
        cli_cnt[cli]+=1; cli_usd[cli]+=get_importe_usd(t,lines)
    top_cli = sorted(cli_usd.items(),key=lambda x:-x[1])[:15]
    cli_rows = "".join(f'<tr><td>{n}</td><td class="r">{cli_cnt[n]}</td><td class="r">{fmt_usd(u)}</td></tr>' for n,u in top_cli)

    # Promedio mensual por influencer (últimos 6 meses)
    ult6 = sorted(by_mes.keys())[-6:]
    tal_mes = defaultdict(lambda: defaultdict(lambda: {"cnt":0,"usd":0.0}))
    for t in tasks:
        fp = get_fecha_pub(t)
        if not fp or (fp.year,fp.month) not in ult6: continue
        tal = parse_talento(t.get("name",""), get_linea(t, lines))
        if not tal: continue
        key = (fp.year,fp.month)
        tal_mes[tal][key]["cnt"]+=1
        tal_mes[tal][key]["usd"]+=get_importe_usd(t,lines)
    top_talentos = [n for n,d in tal_mes.items() if len(d)>=2][:20]
    ph = "".join(f'<th class="r">{mes_label(y,m)}</th>' for y,m in ult6)
    prows = ""
    for tal in top_talentos:
        cells = ""
        for key in ult6:
            d = tal_mes[tal].get(key)
            if d and d["cnt"]:
                prom = d["usd"]/d["cnt"]
                usd_s = fmt_usd(d["usd"]); cnt_s = d["cnt"]
                cells += f'<td class="r" title="{usd_s} / {cnt_s} cont.">{fmt_usd(prom)}</td>'
            else:
                cells += '<td class="r" style="color:var(--tx3)">—</td>'
        prows += f"<tr><td>{tal}</td>{cells}</tr>"

    return f"""<div class="g2">
  <div class="card"><div class="card-h"><div class="dot" style="background:var(--a1)"></div>Contenidos por Mes</div><div class="card-b"><div class="hbars">{hbars}</div></div></div>
  <div class="card"><div class="card-h"><div class="dot" style="background:var(--a4)"></div>Top Talentos Histórico</div><div class="card-b"><div class="tw"><table><thead><tr><th>Talento</th><th class="r">Cont.</th><th class="r">USD acum.</th></tr></thead><tbody>{tal_rows}</tbody></table></div></div></div>
</div>
<div class="card" style="margin-bottom:18px"><div class="card-h"><div class="dot" style="background:var(--a3)"></div>Promedio USD por Contenido (por Influencer)<span class="card-h-note">USD del mes ÷ cantidad de contenidos.</span></div>
  <div class="card-b"><div class="tw"><table><thead><tr><th>Influencer</th>{ph}</tr></thead><tbody>{prows}</tbody></table></div></div></div>
<div class="card"><div class="card-h"><div class="dot" style="background:var(--a5)"></div>Top Clientes Histórico</div>
  <div class="card-b"><div class="tw"><table><thead><tr><th>Cliente</th><th class="r">Cont.</th><th class="r">USD acum.</th></tr></thead><tbody>{cli_rows}</tbody></table></div></div></div>"""

def build_accionables(kpis, lines, pais, tab):
    tag = "[Local]" if tab=="local" else "[Intl]"
    items = []
    venc = kpis["vencidas"]
    if venc:
        mas = max(venc, key=lambda t: (date.today()-get_fecha_est(t)).days if get_fecha_est(t) else 0)
        fe = get_fecha_est(mas)
        dias = (date.today()-fe).days if fe else 0
        items.append(("urgente","🚨","Urgente",
            f'{tag} {len(venc)} campaña(s) vencida(s). Mayor urgencia: {mas.get("name","")[:50]} — {fmt_usd(get_importe_usd(mas,lines))} · {dias}d de atraso.'))
    if kpis["sin_fecha"]>0:
        items.append(("operativo","🔧","Operativo",
            f'{tag} {kpis["sin_fecha"]} subtareas sin fecha estimada. Asignar fechas para mejorar planificación.'))
    if kpis["vel"]>30:
        items.append(("riesgo","⏱","Riesgo Operativo",
            f'{tag} Velocidad de cierre: {kpis["vel"]} días promedio.'))
    if kpis["tasa"]<0.5 and kpis["total_est"]>5:
        items.append(("alerta","📉","Alerta Implementación",
            f'{tag} Tasa: {kpis["tasa"]:.0%}. Solo {kpis["cnt_mes"]} de {kpis["total_est"]} campañas publicadas.'))
    if not items:
        items.append(("ok","✅","Sin alertas críticas",f'{tag} No se detectaron alertas para este período.'))
    html = '<div class="acc">'
    for cls,ico,lbl,txt in items:
        html += f'<div class="ai ai-{cls}"><span class="aico">{ico}</span><div><div class="albl">{lbl}</div><div class="atxt">{txt}</div></div></div>'
    return html + '</div>'

def build_tab(pid, prefix, tab_cls, sec_label, tasks, lines, budget_pais, pais, tab, mes, anio, active):
    kpis = compute_kpis(tasks, lines, mes, anio)
    active_cls = "active" if active else ""
    impl_html = build_implementado(tasks, lines, kpis, mes, anio)
    bg_html   = build_presupuesto(tasks, lines, budget_pais, mes, anio, tab)
    pipe_html = build_pipeline(tasks, lines, mes, anio)
    hist_html = build_historial(tasks, lines, mes, anio)
    acc_html  = build_accionables(kpis, lines, pais, tab)
    venc_cnt  = len(kpis["vencidas"])
    badge     = f' <span class="badge-r">{venc_cnt}</span>' if venc_cnt else ""
    cnt_total = len(tasks)
    usd_total = sum(get_importe_usd(t,lines) for t in tasks if get_fecha_pub(t))
    tc_ars_str = f"{TC_ARS:.0f}"
    fxnote = f'<div class="fxnote">💱 <strong>Todos los importes en USD.</strong> TC MEP: {tc_ars_str} ARS/USD · BCCh 940 CLP/USD · BanRep ~4.250 COP/USD · BCRP ~3.72 PEN/USD</div>'
    return f"""
<div class="mpanel {active_cls}" id="{pid}_{tab_cls}">
  {fxnote}
  <div class="sec-label {tab_cls}">{sec_label}</div>
  <div class="stabs" id="{prefix}_nav">
    <div class="stab active" onclick="sst('{prefix}','impl')">📊 Implementado</div>
    <div class="stab" onclick="sst('{prefix}','bg')">🎯 Presupuesto</div>
    <div class="stab" onclick="sst('{prefix}','pipe')">⏳ Pipeline</div>
    <div class="stab" onclick="sst('{prefix}','hist')">📅 Historial</div>
    <div class="stab" onclick="sst('{prefix}','acc')">⚡ Accionables{badge}</div>
  </div>
  <div class="spanel active" id="{prefix}_impl">{impl_html}</div>
  <div class="spanel" id="{prefix}_bg">{bg_html}</div>
  <div class="spanel" id="{prefix}_pipe">{pipe_html}</div>
  <div class="spanel" id="{prefix}_hist">{hist_html}</div>
  <div class="spanel" id="{prefix}_acc">{acc_html}</div>
</div>"""

def generate_html(pais, local_tasks, intl_tasks, lines, budget, mes, anio, classified=None, objetivos=None):
    now_str = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M hs ARG")
    label   = PAIS_LABEL[pais]
    pid     = f"campanas_{pais}"
    cnt_l   = len(local_tasks)
    cnt_i   = len(intl_tasks)
    usd_l   = sum(get_importe_usd(t,lines) for t in local_tasks if get_fecha_pub(t))
    usd_i   = sum(get_importe_usd(t,lines) for t in intl_tasks  if get_fecha_pub(t))
    bpais   = budget.get(pais, {})

    tab_local = build_tab(pid, f"{pid}_L", "local",
        f"🏠 Ventas Locales — {label}",
        local_tasks, lines, bpais, pais, "local", mes, anio, True)
    tab_intl  = build_tab(pid, f"{pid}_I", "intl",
        f"🌐 Ventas Internacionales — {label}",
        intl_tasks, lines, bpais, pais, "intl", mes, anio, False)

    tiene_objetivos = pais in PAIS_CODE_MAP.values() and classified is not None and objetivos is not None
    mtab_objetivos = ""
    panel_objetivos = ""
    if tiene_objetivos:
        objetivos_pais = objetivos.get(pais, [])
        obj_html = build_objetivos_talentos(pais, anio, classified, lines, objetivos_pais)
        mtab_objetivos = f"""  <div class="mtab objetivos" onclick="sw('{pid}','objetivos')">
    🎯 Objetivos Talentos &nbsp;<span style="opacity:.6;font-weight:400;font-size:12px">{len(objetivos_pais)} talentos</span>
  </div>
"""
        panel_objetivos = f"""<div class="mpanel" id="{pid}_objetivos">
  <div class="sec-label local">🎯 Objetivos {anio} por Talento — {label}</div>
  {obj_html}
</div>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ZAS | {label}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<div class="hdr">
  <div class="hdr-logo">ZAS</div><div class="hdr-sep">/</div>
  <div class="hdr-title">Dashboard de Implementaciones</div>
  <div class="hdr-badge">{label}</div>
  <div class="hdr-date">{now_str}</div>
</div>
<div id="pg_{pid}">
<div class="mtabs">
  <div class="mtab local active" onclick="sw('{pid}','local')">
    🏠 Local &nbsp;<span style="opacity:.6;font-weight:400;font-size:12px">{cnt_l} cont · {fmt_usd(usd_l)}</span>
  </div>
  <div class="mtab intl" onclick="sw('{pid}','intl')">
    🌐 Internacional &nbsp;<span style="opacity:.6;font-weight:400;font-size:12px">{cnt_i} cont · {fmt_usd(usd_i)}</span>
  </div>
{mtab_objetivos}</div>
{tab_local}
{tab_intl}
{panel_objetivos}
</div>
<script>{JS}</script>
</body>
</html>"""


# ─── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    auto = "--auto" in sys.argv
    now  = datetime.now()
    mes  = now.month
    anio = now.year

    print(f"\n{'='*55}")
    print(f"  ZAS Dashboards — {now.strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*55}\n")

    print(f'  TC MEP ARS/USD: {TC_ARS:.0f}')
    odoo_auth()
    tasks, orders, lines = download_data()
    classified = classify(tasks, orders)

    print(f"  Cargando presupuesto...")
    budget = load_budget()

    print(f"  Cargando objetivos de talentos...")
    objetivos = load_objetivos_talentos()
    for p, lst in objetivos.items():
        print(f"    {p}: {len(lst)} talentos con objetivo")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_map = {"argentina":"argentina.html","chile":"chile.html",
                  "colombia":"colombia.html","usa":"usa.html",
                  "peru":"peru.html","internacional":"internacional.html",
                  "mexico":"mexico.html"}

    print(f"\n  Generando HTMLs...")
    for pais, fname in output_map.items():
        lt = classified[pais]['local']
        it = classified[pais]['intl']
        print(f"    {pais}: {len(lt)} local / {len(it)} intl")
        html = generate_html(pais, lt, it, lines, budget, mes, anio, classified, objetivos)
        path = os.path.join(OUTPUT_DIR, fname)
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)

    # Index
    items = "".join(f'<li><a href="{fname}">{pais.upper()}</a></li>' for pais,fname in output_map.items())
    with open(os.path.join(OUTPUT_DIR,"index.html"),"w") as f:
        f.write(f'<!DOCTYPE html><html><head><meta charset="UTF-8"><title>ZAS</title><style>body{{background:#0d0f14;color:#e8eaf0;font-family:Inter,sans-serif;padding:40px}}a{{color:#4f8ef7}}ul{{list-style:none;padding:0}}li{{margin:10px 0;font-size:18px}}</style></head><body><h1 style="color:#4f8ef7">ZAS Dashboards</h1><p style="color:#8b93a8">Actualizado: {now.strftime("%d/%m/%Y %H:%M")}</p><ul>{items}</ul></body></html>')

    print(f"\n  ✅ Listo.")
    if not auto:
        input("  Presioná Enter para salir...")

if __name__ == "__main__":
    main()
